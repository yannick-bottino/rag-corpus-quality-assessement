# src/cqg/report.py
from collections import defaultdict
from .models import CriterionScore, DocScore
from .registry.loader import Registry

def _level(pct: float) -> str:
    if pct >= 90: return "Excellent"
    if pct >= 70: return "Acceptable"
    if pct >= 50: return "Insuffisant"
    return "Inadapté"

def compute_doc_score(doc_id: str, criteria: list[CriterionScore], reg: Registry,
                      parse_confidence: float, config_hash: str,
                      coverage_flag_below: float = 0.7) -> DocScore:
    by_id = {c.id: c for c in reg.criteria}
    num = defaultdict(float); den = defaultdict(float)
    for cs in criteria:
        if cs.status != "scored" or cs.score is None:
            continue
        crit = by_id.get(cs.id)
        if crit is None:  # critere hors registre : ignore plutot que crasher le batch
            continue
        dim = crit.dimension
        num[dim] += cs.weight * cs.score
        den[dim] += cs.weight * reg.scale_max
    dims = {d: round(num[d] / den[d] * 100, 1) for d in den if den[d] > 0}
    gnum = sum(reg.dimension_weights[d] * v for d, v in dims.items())
    gden = sum(reg.dimension_weights[d] for d in dims)
    global_pct = round(gnum / gden, 1) if gden else 0.0
    # Couverture coherente avec le score : un "scored" sans note (score None) ne compte pas.
    scored = sum(1 for c in criteria if c.status == "scored" and c.score is not None)
    not_eval = sum(1 for c in criteria if c.status == "not_evaluated")
    # denom nul (tout na ou liste vide) : couverture indefinie, on retourne 100.0 par convention.
    coverage = round(scored / (scored + not_eval) * 100, 1) if (scored + not_eval) else 100.0
    flags = []
    if coverage < coverage_flag_below * 100:
        flags.append("low_coverage")
    if parse_confidence < 0.5:
        flags.append("low_parse_confidence")
    return DocScore(doc_id=doc_id, global_pct=global_pct, level=_level(global_pct),
                    coverage_pct=coverage, dimensions=dims, criteria=criteria,
                    worst_sections=[], flags=flags, config_hash=config_hash)

# --- Export functions ---
import csv
import json
from pathlib import Path
from openpyxl import Workbook

def write_doc_json(ds: DocScore, out_dir: str) -> str:
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    path = Path(out_dir) / f"{ds.doc_id}.score.json"
    path.write_text(ds.model_dump_json(indent=2), encoding="utf-8")
    return str(path)

def _csv(path: Path, rows: list[list], header: list[str]) -> None:
    # csv.writer echappe les valeurs contenant ; guillemets ou retours ligne (integrite des colonnes).
    # utf-8-sig (BOM) pour un rendu correct des accents a l'ouverture dans Excel.
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(header)
        for r in rows:
            writer.writerow(["" if v is None else v for v in r])

def write_corpus_report(scores: list[DocScore], reg: Registry, out_dir: str) -> dict:
    out = Path(out_dir); out.mkdir(parents=True, exist_ok=True)
    label = {c.id: c.label for c in reg.criteria}
    wb = Workbook()
    ws = wb.active; ws.title = "Synthese"
    ws.append(["doc_id", "score_global_%", "niveau", "couverture_%", "flags"])
    syn_rows = []
    for ds in scores:
        row = [ds.doc_id, ds.global_pct, ds.level, ds.coverage_pct, "|".join(ds.flags)]
        ws.append(row); syn_rows.append(row)
    wd = wb.create_sheet("Detail")
    wd.append(["doc_id", "critere_id", "critere", "tag", "poids", "status", "score", "justification", "preuve"])
    det_rows, rem_rows = [], []
    for ds in scores:
        for c in ds.criteria:
            drow = [ds.doc_id, c.id, label.get(c.id, ""), c.tag, c.weight,
                    c.status, c.score, c.justification, c.evidence]
            wd.append(drow); det_rows.append(drow)
            if c.status == "scored" and c.score is not None and c.score <= 2:
                rem_rows.append([ds.doc_id, c.id, label.get(c.id, ""), c.weight, c.score])
    wr = wb.create_sheet("Remediation")
    wr.append(["doc_id", "critere_id", "critere", "poids", "score_actuel"])
    for r in sorted(rem_rows, key=lambda x: (x[0], -x[3])):
        wr.append(r)
    xlsx = out / "corpus_report.xlsx"; wb.save(xlsx)
    _csv(out / "synthese.csv", syn_rows, ["doc_id", "score_global_%", "niveau", "couverture_%", "flags"])
    _csv(out / "detail.csv", det_rows,
         ["doc_id", "critere_id", "critere", "tag", "poids", "status", "score", "justification", "preuve"])
    _csv(out / "remediation.csv", sorted(rem_rows, key=lambda x: (x[0], -x[3])),
         ["doc_id", "critere_id", "critere", "poids", "score_actuel"])
    return {"xlsx": str(xlsx), "csv_synthese": str(out / "synthese.csv"),
            "csv_detail": str(out / "detail.csv"), "csv_remediation": str(out / "remediation.csv")}
